import itertools
import time
import numpy as np
import pandas as pd
import statsmodels.api as sm
from sklearn import linear_model
from sklearn.metrics import mean_squared_error
from openpyxl import load_workbook

dataset = pd.read_excel('BusinessValuation2.xlsx')
Rdata = pd.DataFrame(dataset)
print(Rdata.head())


# Return R-squared and Adjusted R-squared
def fit_linear_reg(x, y):
    # Fit linear regression model and return RSS and R squared values
    model_k = linear_model.LinearRegression(fit_intercept=True)
    model_k.fit(x, y)
    mse = mean_squared_error(y, model_k.predict(x)) * len(y)
    r_squared = model_k.score(x, y)
    n = len(y)
    p = x.shape[1]
    adj_r_squared = 1 - ((1 - r_squared) * (n - 1)) / (n - p - 1)
    return mse, r_squared, adj_r_squared


# Initialization variables
y = Rdata['P/B Ratio']
x = Rdata.iloc[:, 5:]
print(x)
k = 7  # number of independent variables
MSE_list, R_squared_list, adj_r_squared_list, feature_list = [], [], [], []
numb_indvar = []

# Looping over k = 1 to k = 7 features in X
for k in range(1, len(x.columns) + 1):

    # Looping over all possible combinations: from 11 choose k
    for combo in itertools.combinations(x.columns,k):
        tmp_result = fit_linear_reg(x[list(combo)],y)  # Store temp result
        MSE_list.append(tmp_result[0])
        R_squared_list.append(tmp_result[1])
        adj_r_squared_list.append(tmp_result[2])
        feature_list.append(combo)
        numb_indvar.append(len(combo))

# Store in DataFrame
bess = pd.DataFrame({'numb_indvar': numb_indvar,'MSE': MSE_list, 'R_squared':R_squared_list, 'Adjusted_R_squared':adj_r_squared_list, 'features':feature_list})
# replace indices

# Add a Series of Mallow's Cp Statistic to the dataframe
no_indvar = 7  # number of independent variables
no_para = 8  # total number of parameters


# Save to another sheet in the same excel file
writer = pd.ExcelWriter('BusinessValuation2.xlsx',engine='openpyxl')
book = load_workbook('BusinessValuation2.xlsx')
writer.book = book
bess.to_excel(writer, sheet_name='BESS')  # Input data into the sheet "Pearson Corr" in excel
writer.save()
print('Saved')

'''
# Create new sheet containing Pearson Correlation
print('all column titles', Rdata.columns)
print('slicing rows', Rdata[4:7])
pearsoncorr = Rdata.corr(method='pearson')
print(pearsoncorr)
writer = pd.ExcelWriter('BusinessValuation2.xlsx',engine='openpyxl')
book = load_workbook('BusinessValuation2.xlsx')
writer.book = book
pearsoncorr.to_excel(writer, sheet_name='Pearson Corr')  # Input data into the sheet "Pearson Corr" in excel
writer.save()
print('Saved')

# pandas tutorial
# Show the columns
print(Rdata['Company'].head())
print(Rdata[['Company', 'D2834']].head())
print(Rdata.Company.head())  # DataFrame.column_title

# Select rows
i1 = Rdata.index.get_loc('APPY')
i2 = Rdata.index.get_loc('ACAD')
print(i1, i2)
print(Rdata.iloc[[i2, i1]])
print(Rdata.loc[['ACAD', 'APPY']])

# Select value
print(Rdata.iat[0, 7])  # 1st row, 8th column
print(Rdata.at['ACAD', 'P/B Ratio'])

# Select rows based Boolean value
print(Rdata.D2834 < 1)
print(Rdata[(Rdata['LN(Assets)'] > 10) & (Rdata['LN(Assets)'] < 12)])  # show columns with LN(Assets) in (10,12) range
'''